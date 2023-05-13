# -*- coding: utf-8 -*-
"""
Created on Sat Sep  4 16:17:05 2021

@author: Sakshi
"""

# choosing possible primers
import pandas as pd
import configparser
import sys
name1 = sys.argv[1]
config = configparser.ConfigParser()
#elimination conditions
config.read_file(open(r'config.txt'))
t1 = config.get('My Section', 'temperature(>)',fallback='No Value Entered')
t2 = config.get('My Section', 'temperature(<)',fallback='No Value Entered')
gc1 = config.get('My Section', 'gc(>)',fallback='No Value Entered')
gc2 = config.get('My Section', 'gc(<)',fallback='No Value Entered')
t1=float(t1)
t2=float(t2)
gc1=float(gc1)
gc2=float(gc2)

                             
import openpyxl
  
# load excel with its path
wrkbk = openpyxl.load_workbook(name1+" MIPs.xlsx")
sh = wrkbk.active

# iterate through excel and display data
elim=[]
good=[]

for i in range(1, sh.max_row+1):
    row=[]
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
        row.append(cell_obj.value)
    #print(row)
    if i == 1:
        continue
    else:
        if float(row[4]) > t1 or float(row[4]) < t2:
            
            elim.append(row)
            
        
        elif float(row[5]) < gc2 or float(row[5]) > gc1:
            elim.append(row)
            
        elif float(row[8]) > t1 or float(row[8]) < t2:
            elim.append(row)
            
        elif float(row[9]) < gc2 or float(row[9]) > gc1:
            elim.append(row)
            
        elif row[6] == 'yes' or row[10] == 'yes':
            elim.append(row)
            
        else:
            good.append(row)
            
        
        
df = pd.DataFrame(elim, columns = ["Def Line","Main Sequence","Target region","Extension Arm","TM 1","GC content 1","Continuous stretch","Ligation Arm","TM 2","GC content 2","Continuous stretch 2","Organism"])
df2 = pd.DataFrame(good, columns = ["Def Line","Main Sequence","Target region","Extension Arm","TM 1","GC content 1","Continuous stretch","Ligation Arm","TM 2","GC content 2","Continuous stretch 2","Organism"])

df.to_excel("Eliminated MIPs.xlsx", index=False)
df2.to_excel("Passable MIPs.xlsx", index=False)